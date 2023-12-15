import os
from tkinter.filedialog import askdirectory, askopenfilename
import zipfile
import openpyxl as xl



from Misc import *


# noinspection PyMethodMayBeStatic
class InvoiceReader:

    DEFAULT_DEST_SHEETNAME = "Sheet"
    DEFAULT_CHRONO_SHEETNAME = "salesInvoice"
    DEFAULT_DEST_WORKBOOKNAME = "hsmapping.xlsx"

    def __init__(self):
        self.data = dict()  # the master data
        self.foldername = ""
        self.count_data = 0

        # automatically call getFoldername and retrieveData upon creation.
        self.getFoldername()
        self.retrieveInvoiceData()

    def __del__(self):
        pass

    def getFoldername(self) -> None:
        """
        Populates the self.foldername by opening a tkinter folderdialog box to select target
        :return: None
        """
        # select folder where operation starts.
        self.foldername = askopenfilename(title="Select invoice zip folder:", defaultextension=".zip")

    def _retrieveInvoiceData(self) -> None:
        """
        Internal mechanics of the retrieval of data.
        Will populate self.data from a folder containing all the invoices in .xlsx format.
        It boldly assumes that the folder only contains .xlsx files in the Chronopost Invoice template fashion.
        :return: None
        """

        # checking if a folder has been selected first.
        if self.foldername == "":
            log("Select a folder first...", LogLevel.WARNING)
            return

        log("Retrieving data...")

        # important to reset attributes / class members in case error occured.
        self.data.clear()
        self.count_data = 0

        with zipfile.ZipFile(self.foldername, 'r') as zip_ref:
            zip_ref.extractall(self.foldername[:-4])

        # Goes through each file in the folder (assuming that all are Chronopost Invoices of same format)
        # Then, reads all the descriptions and save them in the global self.data being a dictionary with
        # waybill as the key.
        for path, dirs, files in os.walk(self.foldername[:-4]):
            for file in files:
                wb = xl.load_workbook(os.path.join(path, file))
                sh = wb[InvoiceReader.DEFAULT_CHRONO_SHEETNAME]     # Opening the default sheetname used by Chronopost

                # Below will create a new key from waybill and add a tuple of list[name, telNo] &
                # and an empty list which will hold the items details.
                awb = str(sh.cell(row=7, column=2).value)
                self.data[awb] = [str(sh.cell(row=5, column=4).value),
                                  str(sh.cell(row=6, column=4).value)], list()

                x = 10      # first row where we get the shipment description.

                ttval = 0.0

                # Will iterate from the first description line (row 10) till it finds an empty column.
                while sh.cell(row=x, column=2).value not in (None, ""):
                    self.data[awb][1].append(
                        (
                            str(sh.cell(row=x, column=2).value),            # Description
                            eval(str(sh.cell(row=x, column=3).value)),      # Units
                            eval(str(sh.cell(row=x, column=4).value)),      # Unit Price
                            eval(str(sh.cell(row=x, column=5).value))       # Total Price.
                        )
                    )
                    ttval += eval(str(sh.cell(row=x, column=5).value))
                    x += 1      # next line
                    self.count_data += 1
                sh.cell(row=x, column=4).value = "Total Amount (USD)"
                sh.cell(row=x, column=5).value = ttval
                wb.save(os.path.join(path, awb + "_INVOICE_upload-invoice.xlsx"))
                wb.close()      # closing after finishing.

    def retrieveInvoiceData(self) -> None:
        """
        This is a wrapper to ensure that the most common errors are catered for and treated accordingly.
        :return: None
        """
        try:
            self._retrieveInvoiceData()
            log("Data retrieved...")
        except OSError as e:    # very often because Excel is opened by user.
            self.data.clear()       # clearing data to prevent bug from carrying onwards
            log("Couldn't read all excels, close any opened excel and retry...", LogLevel.ERROR)

    def transposeDescriptions(self) -> None:
        """
        Will take the Descriptions found from Memory and will transpose in a new Excel where the user can enter data.
        Will also break down the component into
        :return:
        """
        destinationFolder = askdirectory(title="Select the destination folder:")
        if destinationFolder == "":
            log("No destination folder has been selected...")
            return

        wb = xl.Workbook()
        x = 1
        sh = wb[InvoiceReader.DEFAULT_DEST_SHEETNAME]
        descArray = self.categorizeDescriptions()   # will sort the values in a better order for convenience.

        for desc in descArray:
            sh.cell(row=x, column=1).value = desc
            x += 1
        wb.save(os.path.join(destinationFolder, InvoiceReader.DEFAULT_DEST_WORKBOOKNAME))

    def categorizeDescriptions(self) -> list:
        """
        It will take the existing self.data and will try to return a list that has been sorted into category
        to ease the identification process.
        :return:
        """
        if len(self.data.keys()) == 0:      # means the data dict is still empty.
            return list(self.data)   # ensuring to return a list even if empty.

        descSet = set()
        for key in self.data.keys():
            for elem in self.data[key][1]:      # means it is iterating each item from each waybills entered in dict.
                descSet.add(elem[0])

        # Our category bins to work with.
        hat = list()
        bags = list()
        shoes = list()
        jewels = list()
        garment = list()
        other = list()

        # Below will iterate values from the set of description (descSet)
        for elem in descSet:
            stripelem = elem.lower().replace(' ', '')       # stripping values only once for performance.

            # below attempts to categorize from the existing words in the description.
            if stripelem.find("hat") > -1:
                hat.append(elem)
            elif stripelem.find("shoe") > -1 or stripelem.find("sandal") > -1 or stripelem.find("slipper") > -1:
                shoes.append(elem)
            elif stripelem.find("bag") > -1 or stripelem.find("backpack") > -1:
                bags.append(elem)
            elif stripelem.find("ring") > -1 or stripelem.find("necklace") > -1:
                jewels.append(elem)
            elif (stripelem.find("%") > -1 and stripelem.find("poly") > -1) or stripelem.find("cotton") > -1 or \
                    stripelem.find("cloth") > -1 or stripelem.find("fiber") > -1:
                garment.append(elem)
            else:
                other.append(elem)

        # join everything into a single list to then return
        hat.extend(bags)
        hat.extend(shoes)
        hat.extend(jewels)
        hat.extend(garment)
        hat.extend(other)

        return hat


class EndNarrator:

    def __init__(self):
        self.wb = xl.Workbook()
        self.sh = self.wb.worksheets[0]

    def __del__(self):
        self.wb.save(os.path.join(os.environ["USERPROFILE"], "Downloads/session_log.xlsx"))

    def transposeDeductions(self, custUnfound, unpaired, newItems):
        self.sh.cell(row=1, column=1).value = "Customer Not Found"
        self.sh.cell(row=1, column=2).value = "Invoice/NOA mismatches"
        self.sh.cell(row=1, column=3).value = "New Item Found"
        r = 2
        for elem in custUnfound:
            self.sh.cell(row=r, column=1).value = elem
            r += 1
        r = 2
        for elem in unpaired:
            self.sh.cell(row=r, column=2).value = elem
            r += 1
        r = 2
        for elem in newItems:
            self.sh.cell(row=r, column=3).value = elem
            r += 1


if __name__ == '__main__':
    pass
